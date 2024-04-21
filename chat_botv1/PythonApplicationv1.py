# -*- coding: cp1251 -*-
import telebot
from gigachat import GigaChat
from similar_text import similar_text
import pandas as pd
from similar_text import similar_text
from dateutil import parser
from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta
import dateparser

def extract_relative_date_from_text(text):
    date = datetime(2024, 4, 3)
    # Используем библиотеку dateparser для извлечения относительной даты
    extracted_date = dateparser.parse(text, settings={"RELATIVE_BASE": date})
    return extracted_date

def classify(text):
    input_text = text
    with open('prompt.txt', 'r',encoding="utf-8") as file:
        text = file.read()
    #print(text)
    with GigaChat(credentials='MmQyNzJmNzAtNTRhMi00ZTg5LTgxNGMtMWQ2OTJhNzg0MmFmOmU5MTY0YjlmLTVhODMtNGI2Yi1iZTU5LTJkYzM5ZjJiNjk3YQ==', verify_ssl_certs=False) as giga:
        #response_text = text + input_text
        response_text="Данный текст содержит дату в произвольном виде (например '1 марта', или 'позавчера'), выдели дату из текста, считая что сегодня 03.04.2024 и выведи только дату в формате '<дд.мм.гггг>, текст: '"+input_text
        response = giga.chat(response_text)
        date1 = response.choices[0].message.content
    print(date1)
    candidate_labels = ['Б.Рокоссовского', 'Черкизовская', 'Преображенск. пл', 'Сокольники СЛ', 'Красносельская', 'Комсомольск. СЛ', 'Красные ворота', 'Чистые пруды', 'Лубянка',
    'Охотный ряд', 'Кропоткинская', 'Парк культуры СЛ', 'Фрунзенская', 'Спортивная', 'Воробьёвы горы', 'Университет', 'Пр-т Вернадск.СЛ', 'Юго-западная',
    'Красногвардейск.', 'Домодедовская', 'Орехово', 'Царицыно', 'Кантемировская', 'Коломенская', 'Автозаводская', 'Павелецкая ЗЛ', 'Новокузнецкая',
    'Театральная(Зам)', 'Тверская', 'Маяковская', 'Белорусская ЗЛ', 'Динамо', 'Аэропорт', 'Сокол', 'Войковская', 'Водный Стадион', 'Речной вокзал',
    'Каширская', 'Каховская', 'Варшавская', 'Киевская АПЛ', 'Смоленская АПЛ', 'Арбатская АПЛ', 'Пл. Революции', 'Курская АПЛ', 'Бауманская', 'Электрозав-я АПЛ',
    'Семёновская', 'Партизанская', 'Измайловская', 'Первомайская', 'Щёлковская', 'Александр. сад', 'Арбатская ФЛ', 'Смоленская ФЛ', 'Студенческая',
    'Кутузовская ФЛ', 'Фили', 'Багратионовская', 'Филёвский парк', 'Пионерская', 'Кунцевская ФЛ', 'Молодёжная', 'Крылатское', 'Белорусская КЛ',
    'Новослободская', 'Проспект Мира КЛ', 'Комсомольская КЛ', 'Курская КЛ', 'Таганская КЛ', 'Павелецкая КЛ', 'Добрынинская', 'Октябрьская КЛ',
    'Парк культуры КЛ', 'Киевская КЛ', 'Краснопресненск.', 'Новогиреево', 'Перово', 'Ш. Энтузиастов', 'Авиамотор-я КалЛ', 'Площадь Ильича',
    'Марксистская', 'Третьяковск.КалЛ', 'Медведково', 'Бабушкинская', 'Свиблово', 'Ботанический сад', 'ВДНХ', 'Алексеевская', 'Рижская',
    'Пр-кт Мира КРЛ', 'Сухаревская', 'Тургеневская', 'Китай-город КРЛ', 'Третьяковск. КРЛ', 'Октябрьская КРЛ', 'Шаболовская', 'Ленинский пр-т',
    'Академическая', 'Профсоюзная', 'Новые Черёмушки', 'Калужская', 'Беляево', 'Коньково', 'Тёплый стан', 'Ясенево', 'Новоясеневская',
    'Выхино', 'Рязанский пр-т', 'Кузьминки', 'Текстильщики ТКЛ', 'Волгоградский пр', 'Пролетарская', 'Таганская ТКЛ', 'Китай-город(Т-К)',
    'Кузнецкий мост', 'Пушкинская', 'Баррикадная', 'Улица 1905 года', 'Беговая', 'Полежаевская', 'Октябрьское поле', 'Щукинская', 'Тушинская',
    'Сходненская', 'Планерная', 'Алтуфьево', 'Бибирево', 'Отрадное', 'Владыкино', 'Петр.-Разумовск.', 'Тимирязевская', 'Дмитровская', 'Савёловская СТЛ',
    'Менделеевская', 'Цветной Бульвар', 'Чеховская', 'Боровицкая', 'Полянка', 'Серпуховская', 'Тульская', 'Нагатинская', 'Нагорная', 'Нахимовский пр-т',
    'Севастопольская', 'Чертановская', 'Южная', 'Пражская', 'Чкаловская', 'Римская', 'Крест. застава', 'Кожуховская', 'Печатники ЛДЛ', 'Волжская', 'Люблино',
    'Братиславская', 'Марьино', 'Стенд', 'Библ. им. Ленина', 'Дубровка', 'Ул. Ак. Янгеля', 'Аннино', 'Административная', 'Б-р Дм. Донского',
    'Парк Победы АПЛ', 'Старокачаловская', 'Ул. Скобелевская', 'Б-р Адм. Ушакова', 'Ул. Горчакова', 'Бунинская аллея', 'Трубная', 'Кунцевская АПЛ',
    'Строгино', 'Каширская (Ках.)', 'Киевская (Фил.)', 'Сретенский б-р', 'Славянский бульв', 'Мякинино', 'Волоколамская', 'Митино', 'Достоевская',
    'Марьина Роща ЛДЛ', 'Борисово', 'Шипиловская', 'Зябликово', 'Новокосино', 'Алма-Атинская', 'Пятницкое шоссе', 'Жулебино', 'Лермонтовский пр',
    'Котельники', 'Битцевский парк', 'Лесопарковая', 'Дел. центр СолЛ', 'Парк Победы КС', 'Спартак', 'Тропарёво', 'Румянцево', 'Саларьево', 'Технопарк',
    'Шелепиха МЦК', 'Шоссе энтуз.МЦК', 'Панфиловская МЦК', 'Стрешнево МЦК', 'Балтийская МЦК', 'Коптево МЦК', 'Лихоборы МЦК', 'Владыкино МЦК',
    'Окружная МЦК', 'Б-р Рокоссов.МЦК', 'Локомотив МЦК', 'Измайлово МЦК', 'Сокол. гора МЦК', 'Андроновка МЦК', 'Угрешская МЦК', 'Дубровка МЦК',
    'Автозавод. МЦК', 'ЗИЛ МЦК', 'Хорошёво МЦК', 'Ростокино МЦК', 'Ботан. сад МЦК', 'Белокаменная МЦК', 'Новохохлов-я МЦК', 'Верхн. Котлы МЦК',
    'Крымская МЦК', 'Нижегородск. МЦК', 'Лужники МЦК', 'Кутузовская МЦК', 'Москва-Сити МЦК', 'Зорге МЦК', 'Пл. Гагарина МЦК', 'Бутырская', 'Фонвизинская',
    'Петровско-Разум.', 'Минская', 'Ломоносов-ий п-т', 'Раменки', 'Ломоносовск.пр-т', 'Парк Победы СолЛ', 'Воробьевы горы С', 'Комсомольская 2',
     'Выхино 2', 'Петровский парк', 'Дел. центр БКЛ', 'Ховрино', 'Шелепиха', 'Хорошёвская', 'ЦСКА', 'Воробьёвы горы С', 'Селигерская',
     'Верхние Лихоборы', 'Окружная', 'Шелепиха(Солнц.)', 'Хорошев-я(Солнц)', 'ЦСКА(Солнц.)', 'Петр.парк(Солнц)', 'Мичурин.пр-тСолЛ', 'Озёрная',
     'Говорово', 'Солнцево', 'Боровское шоссе', 'Новопеределкино', 'Рассказовка', 'Беломорская', 'Савёловская БКЛ', 'Филатов луг', 'Прокшино', 'Ольховая',
     'Коммунарка', 'Косино', 'Ул.Дмитриевского', 'Лухмановская', 'Некрасовка', 'Юго-Восточная', 'Окская', 'Стахановская', 'Нижегородская',
     'Авиамоторная БКЛ', 'Лефортово', 'ТПУ Рязанская', 'Авиамоторная нек', 'Лефортово нек.', 'Лефортово БКЛ', 'Электрозав-я БКЛ', 'Народное Ополч-е',
                        'Мнёвники', 'Электрозав-я нек', 'Хорошёвская(Мн)', 'Зюзино',
    'Воронцовская', 'Новаторская', 'Пр-т Вернад. БКЛ', 'Мичурин.пр-т БКЛ', 'Аминьевская', 'Давыдково', 'Кунцевская БКЛ', 'Терехово', 'Марьина Роща БКЛ',
    'Рижская БКЛ', 'Сокольники БКЛ', 'Текстильщики БКЛ', 'Печатники БКЛ', 'Нагатинский З-н', 'Кленовый бульвар',
    'Нижегород-я БКЛ', 'Каширская (Зам)', 'Нижегород-я НБС', 'К', 'Пыхтино', 'Аэропорт Внуково', 'Яхромская', 'Лианозово', 'Физтех', 'Текстильщики СЦ']
    
    with GigaChat(credentials='MmQyNzJmNzAtNTRhMi00ZTg5LTgxNGMtMWQ2OTJhNzg0MmFmOmU5MTY0YjlmLTVhODMtNGI2Yi1iZTU5LTJkYzM5ZjJiNjk3YQ==', verify_ssl_certs=False) as giga:
        #response_text = ', '.join(map(str,candidate_labels)) + 'выведи ту станцию из списка, которая упоминается в следующем предложении' + input_text
        response_text = 'выведи ту станцию метро Москвы, которая упоминается в следующем предложении, пиши только название станции в формате "<Название станции>": ' + input_text
        response = giga.chat(response_text)
        date2 = response.choices[0].message.content
   
    print(date2)
    l = list(map(similar_text,[date2.lower() for i in range(len(candidate_labels))],candidate_labels))
    res = list()
    for i in range(len(candidate_labels)):
      res.append([l[i],candidate_labels[i]])
    res1 = sorted(res,reverse=True)[0][1]

    try:
        res = bool(parser.parse(date1))
    except ValueError:
        try:
            date1 = extract_relative_date_from_text(date1).date().strftime('%d.%m.%Y')
        except:
            date1="03.04.2024"

    return (date1,res1)


import openpyxl
from datetime import datetime

def is_date_column(sheet, col_num, dates):
    cell_value = sheet.cell(row=1, column=col_num).value
    if isinstance(cell_value, datetime):
        date_str = cell_value.strftime('%d.%m.%Y')
        if date_str in dates:
            return True
    return False

def get_sum_for_station_and_dates(file_path, station_name, dates):
    # Открываем файл Excel
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Получаем индексы столбцов для станции и дат
    station_column = None
    date_columns = []
    for col_num in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col_num).value == "Станция":
            station_column = col_num
        elif is_date_column(sheet, col_num, dates):
            date_columns.append(col_num)

    if station_column is None:
        return "Столбец 'Станция' не найден в файле"
    
    if not date_columns:
        return "Ни однa из дат указанных в массиве не найдена в файле"

    # Ищем строки, соответствующие указанной станции
    station_rows = []
    for row_num in range(2, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=station_column).value == station_name:
            station_rows.append(row_num)

    if not station_rows:
        return f"Станция '{station_name}' не найдена в файле"

    # Суммируем значения для каждой даты
    total_sum = 0
    for row_num in station_rows:
        for date_column in date_columns:
            value = sheet.cell(row=row_num, column=date_column).value
            if value is not None:
                total_sum += value

    return total_sum

# Пример использования
file_path = "./test.xlsx"
#station_name = "Б.Рокоссовского"
#dates = ["01.04.2024", "02.04.2024", "03.04.2024"]

#result = get_sum_for_station_and_dates(file_path, station_name, dates)
#print("Сумма значений для указанной станции и дат:", result)


import speech_recognition as sr
from pydub import AudioSegment
import os

def recognize_speech(file_path):
    # Преобразуем файл OGG в WAV
    sound = AudioSegment.from_ogg(file_path)
    wav_path = "temp.wav"
    sound.export(wav_path, format="wav")

    recognizer = sr.Recognizer()

    # Открываем WAV файл
    with sr.AudioFile(wav_path) as source:
        audio_data = recognizer.record(source)  # Считываем аудиоданные

    try:
        # Проводим распознавание речи
        text = recognizer.recognize_google(audio_data, language="ru-RU")
        return text
    except sr.UnknownValueError:
        return "Не удалось распознать речь"
    except sr.RequestError as e:
        return f"Ошибка при отправке запроса к сервису распознавания: {e}"
    finally:
        # Удаляем временный WAV файл
        os.remove(wav_path)


bot = telebot.TeleBot('6983581621:AAF7YZoFUBaEhM-t2HVsg1qetD9Mn95ppOk')

@bot.message_handler()
def save_message(message):
    #with open("messages.txt", "a", encoding="utf-8") as file:
        #file.write(message.text + "\n")
    print(f'Запрос: {message.text}')
    res=classify(message.text)
    print(f'Ответ: {res}')
    bot.reply_to(message, f'Распознанные токены: {res}')
    result = get_sum_for_station_and_dates(file_path, res[1], res[0])
    bot.reply_to(message, f'Результат: {result}')
    print(f'Ответ: {result}')

@bot.message_handler(content_types=['voice'])
def voice_processing(message):
    file_info = bot.get_file(message.voice.file_id)
    downloaded_file = bot.download_file(file_info.file_path)
    with open('new_file.ogg', 'wb+') as new_file:
        new_file.write(downloaded_file)
    recognized_text = recognize_speech("./new_file.ogg")
    print("Распознанный текст: ", recognized_text)
    res=classify(recognized_text)
    print(f'Ответ: {res}')
    bot.reply_to(message, f'Распознанные токены: {res}')
    result = get_sum_for_station_and_dates(file_path, res[1], res[0])
    bot.reply_to(message, f'Результат: {result}')
    print(f'Ответ: {result}')
   
bot.polling(none_stop=True)

